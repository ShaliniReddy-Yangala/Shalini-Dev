from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
import io
import pandas as pd
from datetime import datetime
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from app.Employee_Master_Report.emp_models.dropdowns import (
    Category, ExcludedFromPayroll, MaritalStatus, BloodGroup,
    AddressType, RelationType, TypeOfDegree, JobType, AssetType, AssetStatus, Gender, Title
)
from app.models import Department as DeptModel, Client as ClientModel, GenderDB as CandidateGender
from app.Employee_Master_Report.emp_models.employee_master import EmployeeMaster, AssetHistory

router = APIRouter(prefix="/employee-entry", tags=["Employee Entry - Excel Template"])


@router.get("/download-excel-template")
def download_excel_template(
    db: Session = Depends(get_db),
    upload_type: str = Query(default="create", description="create | update"),
    employee_id: str | None = Query(default=None)
):
    """Download Excel template for employee data entry with all fields and multiple rows for dynamic sections"""
    
    try:
        # Create a BytesIO buffer to hold the Excel file
        output = io.BytesIO()
        
        # Build Lists sheet data from DB
        def col(values):
            return list(dict.fromkeys([v for v in values if v]))  # unique, preserve order

        categories = col([c.type_of_category for c in db.query(Category).order_by(Category.type_of_category.asc()).all()])
        excluded = col([e.value for e in db.query(ExcludedFromPayroll).order_by(ExcludedFromPayroll.value.asc()).all()])
        marital = col([m.status for m in db.query(MaritalStatus).order_by(MaritalStatus.status.asc()).all()])
        blood = col([b.group for b in db.query(BloodGroup).order_by(BloodGroup.group.asc()).all()])
        address_types = col([a.type for a in db.query(AddressType).order_by(AddressType.type.asc()).all()])
        relation_types = col([r.type for r in db.query(RelationType).order_by(RelationType.type.asc()).all()])
        degree_types = col([t.degree for t in db.query(TypeOfDegree).order_by(TypeOfDegree.degree.asc()).all()])
        job_types = col([j.type for j in db.query(JobType).order_by(JobType.type.asc()).all()])
        asset_types = col([a.type for a in db.query(AssetType).order_by(AssetType.type.asc()).all()])
        asset_statuses = col([s.status for s in db.query(AssetStatus).order_by(AssetStatus.status.asc()).all()])
        # Titles from employee Title master
        titles = col([t.title for t in db.query(Title).order_by(Title.title.asc()).all()])
        # Gender from employee gender master; fallback to candidates GenderDB if empty
        genders_emp = col([g.gender for g in db.query(Gender).order_by(Gender.gender.asc()).all()])
        if not genders_emp:
            try:
                genders_emp = col([g.gender for g in db.query(CandidateGender).order_by(CandidateGender.gender.asc()).all()])
            except Exception:
                genders_emp = []

        # Fetch Departments and Clients using ORM models
        try:
            departments = col([d.name for d in db.query(DeptModel).order_by(DeptModel.name.asc()).all()])
        except Exception:
            departments = []
        try:
            clients = col([c.name for c in db.query(ClientModel).order_by(ClientModel.name.asc()).all()])
        except Exception:
            clients = []

        # Create Excel writer with multiple sheets
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Create Lists sheet first
            lists_df = pd.DataFrame({
                'Departments': pd.Series(departments or ['']),
                'Categories': pd.Series(categories or ['']),
                'ExcludedFromPayroll': pd.Series(excluded or ['']),
                'Titles': pd.Series(titles or ['']),
                'Gender': pd.Series(genders_emp or ['']),
                'MaritalStatus': pd.Series(marital or ['']),
                'BloodGroup': pd.Series(blood or ['']),
                'AddressType': pd.Series(address_types or ['']),
                'RelationType': pd.Series(relation_types or ['']),
                'DegreeType': pd.Series(degree_types or ['']),
                'JobType': pd.Series(job_types or ['']),
                'Clients': pd.Series(clients or ['']),
                'AssetType': pd.Series(asset_types or ['']),
                'AssetStatus': pd.Series(asset_statuses or ['']),
            })
            lists_df.to_excel(writer, sheet_name='Lists', index=False)
            
            # Consolidated Sheet: Employee Details (Single row per employee)
            employee_details_data = {
                # Basic Details
                'Employee ID': [''],
                'DOJ (DD-MM-YYYY)': [''],
                'Designation': [''],
                'Department': [''],
                'Manager Name': [''],
                'Official Contact Number': [''],
                'Official Email ID': [''],
                'Category': [''],
                'Excluded from Payroll': [''],
                # Personal Details
                'Title': [''],
                'First Name': [''],
                'Last Name': [''],
                'Full Name (Auto-generated)': [''],
                'Gender': [''],
                'DOB (DD-MM-YYYY)': [''],
                'Marital Status': [''],
                'DOA (DD-MM-YYYY)': [''],
                'Religion': [''],
                'Blood Group': [''],
                'Mobile No': [''],
                # Contract Details
                'Job Type': [''],
                'Contract End Date (DD-MM-YYYY)': [''],
                'Probation End Date (DD-MM-YYYY)': [''],
                # Bank Details
                'Bank Name': [''],
                'Account No': [''],
                'IFSC Code': [''],
                'Type of Account': [''],
                'Branch': [''],
                # Communication Details
                'PAN Card No': [''],
                'Aadhar No': [''],
                'Name as per Aadhar': [''],
                'Passport No': [''],
                'Passport Issued Date (DD-MM-YYYY)': [''],
                'Passport Expiry Date (DD-MM-YYYY)': [''],
                'Personal Email ID': [''],
                'Current UAN No': [''],
                # Salary Details
                'Gross Salary per Month': [''],
                'Tax Regime': [''],
                # Health Insurance
                'Policy No': [''],
                'Commencement Date (DD-MM-YYYY)': [''],
                'End Date (DD-MM-YYYY)': [''],
                'Amount': [''],
                'Covered Members': [''],
                'Duration': [''],
                'Insurer Name': [''],
            }
            # Prefill for update mode
            if upload_type == "update":
                if employee_id:
                    # Single employee update
                    emp = db.query(EmployeeMaster).filter(EmployeeMaster.employee_id == employee_id).first()
                    if emp:
                        employee_details_data = {
                            'Employee ID': [emp.employee_id or ''],
                            'DOJ (DD-MM-YYYY)': [emp.doj.strftime('%d-%m-%Y') if emp.doj else ''],
                            'Designation': [emp.designation or ''],
                            'Department': [emp.department or ''],
                            'Manager Name': [emp.manager_name or ''],
                            'Official Contact Number': [emp.official_no or ''],
                            'Official Email ID': [emp.official_email_id or ''],
                            'Category': [emp.category or ''],
                            'Excluded from Payroll': [emp.excluded_from_payroll or ''],
                            'Title': [emp.title or ''],
                            'First Name': [emp.first_name or ''],
                            'Last Name': [emp.last_name or ''],
                            'Full Name (Auto-generated)': [emp.full_name or ''],
                            'Gender': [emp.gender or ''],
                            'DOB (DD-MM-YYYY)': [emp.dob.strftime('%d-%m-%Y') if emp.dob else ''],
                            'Marital Status': [emp.marital_status or ''],
                            'DOA (DD-MM-YYYY)': [emp.doa.strftime('%d-%m-%Y') if emp.doa else ''],
                            'Religion': [emp.religion or ''],
                            'Blood Group': [emp.blood_group or ''],
                            'Mobile No': [emp.mobile_no or ''],
                            'Job Type': [emp.job_type or ''],
                            'Contract End Date (DD-MM-YYYY)': [emp.contract_end_date.strftime('%d-%m-%Y') if emp.contract_end_date else ''],
                            'Probation End Date (DD-MM-YYYY)': [emp.probation_end_date.strftime('%d-%m-%Y') if emp.probation_end_date else ''],
                            'Bank Name': [emp.bank_name or ''],
                            'Account No': [emp.account_no or ''],
                            'IFSC Code': [emp.ifsc_code or ''],
                            'Type of Account': [emp.type_of_account or ''],
                            'Branch': [emp.branch or ''],
                            'PAN Card No': [emp.pan_card_no or ''],
                            'Aadhar No': [emp.aadhar_no or ''],
                            'Name as per Aadhar': [emp.name_as_per_aadhar or ''],
                            'Passport No': [emp.passport_no or ''],
                            'Passport Issued Date (DD-MM-YYYY)': [emp.issued_date.strftime('%d-%m-%Y') if emp.issued_date else ''],
                            'Passport Expiry Date (DD-MM-YYYY)': [emp.expiry_date.strftime('%d-%m-%Y') if emp.expiry_date else ''],
                            'Personal Email ID': [emp.personal_email_id or ''],
                            'Current UAN No': [emp.current_uan_no or ''],
                            'Gross Salary per Month': [float(emp.gross_salary_per_month) if emp.gross_salary_per_month is not None else ''],
                            'Tax Regime': [emp.tax_regime or ''],
                            'Policy No': [emp.policy_no or ''],
                            'Commencement Date (DD-MM-YYYY)': [emp.commencement_date.strftime('%d-%m-%Y') if emp.commencement_date else ''],
                            'End Date (DD-MM-YYYY)': [emp.end_date.strftime('%d-%m-%Y') if emp.end_date else ''],
                            'Amount': [float(emp.amount) if emp.amount is not None else ''],
                            'Covered Members': [emp.covered_members or ''],
                            'Duration': [emp.duration or ''],
                            'Insurer Name': [emp.insurer_name or ''],
                        }
                else:
                    # All employees update - populate with all existing employees
                    employees = db.query(EmployeeMaster).order_by(EmployeeMaster.employee_id.asc()).all()
                    if employees:
                        employee_details_data = {
                            'Employee ID': [emp.employee_id or '' for emp in employees],
                            'DOJ (DD-MM-YYYY)': [emp.doj.strftime('%d-%m-%Y') if emp.doj else '' for emp in employees],
                            'Designation': [emp.designation or '' for emp in employees],
                            'Department': [emp.department or '' for emp in employees],
                            'Manager Name': [emp.manager_name or '' for emp in employees],
                            'Official Contact Number': [emp.official_no or '' for emp in employees],
                            'Official Email ID': [emp.official_email_id or '' for emp in employees],
                            'Category': [emp.category or '' for emp in employees],
                            'Excluded from Payroll': [emp.excluded_from_payroll or '' for emp in employees],
                            'Title': [emp.title or '' for emp in employees],
                            'First Name': [emp.first_name or '' for emp in employees],
                            'Last Name': [emp.last_name or '' for emp in employees],
                            'Full Name (Auto-generated)': [emp.full_name or '' for emp in employees],
                            'Gender': [emp.gender or '' for emp in employees],
                            'DOB (DD-MM-YYYY)': [emp.dob.strftime('%d-%m-%Y') if emp.dob else '' for emp in employees],
                            'Marital Status': [emp.marital_status or '' for emp in employees],
                            'DOA (DD-MM-YYYY)': [emp.doa.strftime('%d-%m-%Y') if emp.doa else '' for emp in employees],
                            'Religion': [emp.religion or '' for emp in employees],
                            'Blood Group': [emp.blood_group or '' for emp in employees],
                            'Mobile No': [emp.mobile_no or '' for emp in employees],
                            'Job Type': [emp.job_type or '' for emp in employees],
                            'Contract End Date (DD-MM-YYYY)': [emp.contract_end_date.strftime('%d-%m-%Y') if emp.contract_end_date else '' for emp in employees],
                            'Probation End Date (DD-MM-YYYY)': [emp.probation_end_date.strftime('%d-%m-%Y') if emp.probation_end_date else '' for emp in employees],
                            'Bank Name': [emp.bank_name or '' for emp in employees],
                            'Account No': [emp.account_no or '' for emp in employees],
                            'IFSC Code': [emp.ifsc_code or '' for emp in employees],
                            'Type of Account': [emp.type_of_account or '' for emp in employees],
                            'Branch': [emp.branch or '' for emp in employees],
                            'PAN Card No': [emp.pan_card_no or '' for emp in employees],
                            'Aadhar No': [emp.aadhar_no or '' for emp in employees],
                            'Name as per Aadhar': [emp.name_as_per_aadhar or '' for emp in employees],
                            'Passport No': [emp.passport_no or '' for emp in employees],
                            'Passport Issued Date (DD-MM-YYYY)': [emp.issued_date.strftime('%d-%m-%Y') if emp.issued_date else '' for emp in employees],
                            'Passport Expiry Date (DD-MM-YYYY)': [emp.expiry_date.strftime('%d-%m-%Y') if emp.expiry_date else '' for emp in employees],
                            'Personal Email ID': [emp.personal_email_id or '' for emp in employees],
                            'Current UAN No': [emp.current_uan_no or '' for emp in employees],
                            'Gross Salary per Month': [float(emp.gross_salary_per_month) if emp.gross_salary_per_month is not None else '' for emp in employees],
                            'Tax Regime': [emp.tax_regime or '' for emp in employees],
                            'Policy No': [emp.policy_no or '' for emp in employees],
                            'Commencement Date (DD-MM-YYYY)': [emp.commencement_date.strftime('%d-%m-%Y') if emp.commencement_date else '' for emp in employees],
                            'End Date (DD-MM-YYYY)': [emp.end_date.strftime('%d-%m-%Y') if emp.end_date else '' for emp in employees],
                            'Amount': [float(emp.amount) if emp.amount is not None else '' for emp in employees],
                            'Covered Members': [emp.covered_members or '' for emp in employees],
                            'Duration': [emp.duration or '' for emp in employees],
                            'Insurer Name': [emp.insurer_name or '' for emp in employees],
                        }
            employee_details_df = pd.DataFrame(employee_details_data)
            employee_details_df.to_excel(writer, sheet_name='Employee Details', index=False)
            
            # Sheet 3: Address Details (Multiple rows - Permanent, Temporary, Office)
            if upload_type == "update" and not employee_id:
                # All employees update - populate with all existing address data
                from app.Employee_Master_Report.emp_models.employee_master import AddressHistory
                address_records = db.query(AddressHistory).order_by(AddressHistory.employee_id.asc(), AddressHistory.address_type.asc()).all()
                if address_records:
                    address_data = {
                        'Employee ID': [addr.employee_id for addr in address_records],
                        'Address Type': [addr.address_type for addr in address_records],
                        'H.No': [addr.h_no or '' for addr in address_records],
                        'Street': [addr.street or '' for addr in address_records],
                        'Street2': [addr.street2 or '' for addr in address_records],
                        'Landmark': [addr.landmark or '' for addr in address_records],
                        'City': [addr.city or '' for addr in address_records],
                        'State': [addr.state or '' for addr in address_records],
                        'Postal Code': [addr.postal_code or '' for addr in address_records],
                        'Complete Address (Auto-generated)': [addr.complete_address or '' for addr in address_records]
                    }
                else:
                    address_data = {
                        'Employee ID': ['', '', ''],
                        'Address Type': ['Permanent', 'Temporary', 'Office'],
                        'H.No': ['', '', ''],
                        'Street': ['', '', ''],
                        'Street2': ['', '', ''],
                        'Landmark': ['', '', ''],
                        'City': ['', '', ''],
                        'State': ['', '', ''],
                        'Postal Code': ['', '', ''],
                        'Complete Address (Auto-generated)': ['', '', '']
                    }
            else:
                address_data = {
                    'Employee ID': ['', '', ''],
                    'Address Type': ['Permanent', 'Temporary', 'Office'],
                    'H.No': ['', '', ''],
                    'Street': ['', '', ''],
                    'Street2': ['', '', ''],
                    'Landmark': ['', '', ''],
                    'City': ['', '', ''],
                    'State': ['', '', ''],
                    'Postal Code': ['', '', ''],
                    'Complete Address (Auto-generated)': ['', '', '']
                }
            address_df = pd.DataFrame(address_data)
            address_df.to_excel(writer, sheet_name='Address Details', index=False)
            
            # Sheet 4: Family Members Details (Multiple rows)
            if upload_type == "update" and not employee_id:
                # All employees update - populate with all existing family data
                from app.Employee_Master_Report.emp_models.employee_master import FamilyMember
                family_records = db.query(FamilyMember).order_by(FamilyMember.employee_id.asc()).all()
                if family_records:
                    family_data = {
                        'Employee ID': [fam.employee_id for fam in family_records],
                        'Relation Type': [fam.relation_type or '' for fam in family_records],
                        'Name': [fam.name or '' for fam in family_records],
                        'DOB (DD-MM-YYYY)': [fam.dob.strftime('%d-%m-%Y') if fam.dob else '' for fam in family_records],
                        'Aadhar Number': [fam.aadhar_number or '' for fam in family_records],
                        'Dependant (Yes/No)': [fam.dependant or '' for fam in family_records]
                    }
                else:
                    family_data = {
                        'Employee ID': ['', '', ''],
                        'Relation Type': ['', '', ''],
                        'Name': ['', '', ''],
                        'DOB (DD-MM-YYYY)': ['', '', ''],
                        'Aadhar Number': ['', '', ''],
                        'Dependant (Yes/No)': ['', '', '']
                    }
            else:
                family_data = {
                    'Employee ID': ['', '', ''],
                    'Relation Type': ['', '', ''],
                    'Name': ['', '', ''],
                    'DOB (DD-MM-YYYY)': ['', '', ''],
                    'Aadhar Number': ['', '', ''],
                    'Dependant (Yes/No)': ['', '', '']
                }
            family_df = pd.DataFrame(family_data)
            family_df.to_excel(writer, sheet_name='Family Members', index=False)
            
            # Sheet 5: Education Details (Multiple rows)
            if upload_type == "update" and not employee_id:
                # All employees update - populate with all existing education data
                from app.Employee_Master_Report.emp_models.employee_master import EducationHistory
                education_records = db.query(EducationHistory).order_by(EducationHistory.employee_id.asc()).all()
                if education_records:
                    education_data = {
                        'Employee ID': [edu.employee_id for edu in education_records],
                        'Type of Degree': [edu.type_of_degree or '' for edu in education_records],
                        'Course Name': [edu.course_name or '' for edu in education_records],
                        'Completed Month (1-12)': [edu.completed_in_month_year.split('-')[0] if edu.completed_in_month_year and '-' in edu.completed_in_month_year else '' for edu in education_records],
                        'Completed Year': [edu.completed_in_month_year.split('-')[1] if edu.completed_in_month_year and '-' in edu.completed_in_month_year else '' for edu in education_records],
                        'School/College Name': [edu.school_college_name or '' for edu in education_records],
                        'Affiliated from University': [edu.affiliated_university or '' for edu in education_records],
                        'Document File Name (PDF/JPG)': ['' for edu in education_records]
                    }
                else:
                    education_data = {
                        'Employee ID': ['', '', ''],
                        'Type of Degree': ['', '', ''],
                        'Course Name': ['', '', ''],
                        'Completed Month (1-12)': ['', '', ''],
                        'Completed Year': ['', '', ''],
                        'School/College Name': ['', '', ''],
                        'Affiliated from University': ['', '', ''],
                        'Document File Name (PDF/JPG)': ['', '', '']
                    }
            else:
                education_data = {
                    'Employee ID': ['', '', ''],
                    'Type of Degree': ['', '', ''],
                    'Course Name': ['', '', ''],
                    'Completed Month (1-12)': ['', '', ''],
                    'Completed Year': ['', '', ''],
                    'School/College Name': ['', '', ''],
                    'Affiliated from University': ['', '', ''],
                    'Document File Name (PDF/JPG)': ['', '', '']
                }
            education_df = pd.DataFrame(education_data)
            education_df.to_excel(writer, sheet_name='Education Details', index=False)
            
            # Sheet 6: Previous Experience Details (Multiple rows)
            if upload_type == "update" and not employee_id:
                # All employees update - populate with all existing experience data
                from app.Employee_Master_Report.emp_models.employee_master import ExperienceHistory
                experience_records = db.query(ExperienceHistory).order_by(ExperienceHistory.employee_id.asc()).all()
                if experience_records:
                    experience_data = {
                        'Employee ID': [exp.employee_id for exp in experience_records],
                        'Company Name': [exp.company_name or '' for exp in experience_records],
                        'Start Date (DD-MM-YYYY)': [exp.start_date.strftime('%d-%m-%Y') if exp.start_date else '' for exp in experience_records],
                        'End Date (DD-MM-YYYY)': [exp.end_date.strftime('%d-%m-%Y') if exp.end_date else '' for exp in experience_records],
                        'Designation': [exp.designation or '' for exp in experience_records],
                        'Department': [exp.department or '' for exp in experience_records],
                        'Office Email ID': [exp.office_email_id or '' for exp in experience_records],
                        'UAN No': [exp.uan_no or '' for exp in experience_records],
                        'PF No': ['' for exp in experience_records],  # PF No is stored in master table
                        'Company Address': ['' for exp in experience_records],  # Company Address is stored in master table
                        'Reference Details -1': ['' for exp in experience_records],  # Reference details are stored in master table
                        'Reference Details -2': ['' for exp in experience_records]
                    }
                else:
                    experience_data = {
                        'Employee ID': ['', '', ''],
                        'Company Name': ['', '', ''],
                        'Start Date (DD-MM-YYYY)': ['', '', ''],
                        'End Date (DD-MM-YYYY)': ['', '', ''],
                        'Designation': ['', '', ''],
                        'Department': ['', '', ''],
                        'Office Email ID': ['', '', ''],
                        'UAN No': ['', '', ''],
                        'PF No': ['', '', ''],
                        'Company Address': ['', '', ''],
                        'Reference Details -1': ['', '', ''],
                        'Reference Details -2': ['', '', '']
                    }
            else:
                experience_data = {
                    'Employee ID': ['', '', ''],
                    'Company Name': ['', '', ''],
                    'Start Date (DD-MM-YYYY)': ['', '', ''],
                    'End Date (DD-MM-YYYY)': ['', '', ''],
                    'Designation': ['', '', ''],
                    'Department': ['', '', ''],
                    'Office Email ID': ['', '', ''],
                    'UAN No': ['', '', ''],
                    'PF No': ['', '', ''],
                    'Company Address': ['', '', ''],
                    'Reference Details -1': ['', '', ''],
                    'Reference Details -2': ['', '', '']
                }
            experience_df = pd.DataFrame(experience_data)
            experience_df.to_excel(writer, sheet_name='Experience Details', index=False)
            
            # Sheet 10: Nominee Details (Single row)
            nominee_data = {
                'Nominee Name': [''],
                'Address': [''],
                'Relation': [''],
                'Age': [''],
                'Proportion': ['']
            }
            nominee_df = pd.DataFrame(nominee_data)
            nominee_df.to_excel(writer, sheet_name='Nominee Details', index=False)
            
            # Sheet 11: Emergency Contact Details (Two rows)
            emergency_data = {
                'Employee ID': ['', ''],
                'Contact Name': ['Emergency Contact 1', 'Emergency Contact 2'],
                'Relation': ['', ''],
                'Contact Number': ['', '']
            }
            emergency_df = pd.DataFrame(emergency_data)
            emergency_df.to_excel(writer, sheet_name='Emergency Contacts', index=False)
            
            # Sheet 12: Onboarding Details (Single row)
            if upload_type == "update" and not employee_id:
                # All employees update - populate with all existing onboarding data
                from app.Employee_Master_Report.emp_models.employee_master import OnboardingHistory, ClientMaster
                onboarding_records = db.query(OnboardingHistory).order_by(OnboardingHistory.employee_id.asc()).all()
                if onboarding_records:
                    onboarding_data = {
                        'Employee ID': [onb.employee_id for onb in onboarding_records],
                        'Client Name': [db.query(ClientMaster).filter(ClientMaster.client_id == onb.client_id).first().client_name if db.query(ClientMaster).filter(ClientMaster.client_id == onb.client_id).first() else '' for onb in onboarding_records],
                        'Effective Start Date (DD-MM-YYYY)': [onb.effective_start_date.strftime('%d-%m-%Y') if onb.effective_start_date else '' for onb in onboarding_records],
                        'Effective End Date (DD-MM-YYYY)': [onb.effective_end_date.strftime('%d-%m-%Y') if onb.effective_end_date else '' for onb in onboarding_records],
                        'Current Onboarding Status (Active/Withdrawn/On Bench)': [onb.onboarding_status or '' for onb in onboarding_records],
                        'Duration (Auto-calculated)': [onb.duration_calculated or '' for onb in onboarding_records],
                        'SPOC': [onb.spoc or '' for onb in onboarding_records],
                        'Department': [onb.onboarding_department or '' for onb in onboarding_records],
                        'Manager': [onb.assigned_manager or '' for onb in onboarding_records]
                    }
                else:
                    onboarding_data = {
                        'Employee ID': [''],
                        'Client Name': [''],
                        'Effective Start Date (DD-MM-YYYY)': [''],
                        'Effective End Date (DD-MM-YYYY)': [''],
                        'Current Onboarding Status (Active/Withdrawn/On Bench)': [''],
                        'Duration (Auto-calculated)': [''],
                        'SPOC': [''],
                        'Department': [''],
                        'Manager': ['']
                    }
            else:
                onboarding_data = {
                    'Employee ID': [''],
                    'Client Name': [''],
                    'Effective Start Date (DD-MM-YYYY)': [''],
                    'Effective End Date (DD-MM-YYYY)': [''],
                    'Current Onboarding Status (Active/Withdrawn/On Bench)': [''],
                    'Duration (Auto-calculated)': [''],
                    'SPOC': [''],
                    'Department': [''],
                    'Manager': ['']
                }
            onboarding_df = pd.DataFrame(onboarding_data)
            onboarding_df.to_excel(writer, sheet_name='Onboarding Details', index=False)
            
            # Sheet 13: Asset Details (Single row)
            asset_data = {
                'Employee ID': [''],
                'Asset Type': [''],
                'Asset Number': [''],
                'Issued Date (DD-MM-YYYY)': [''],
                'Status': ['']
            }
            # Prefill Asset list in update mode
            if upload_type == "update":
                if employee_id:
                    # Single employee update
                    assets = db.query(AssetHistory).filter(AssetHistory.employee_id == employee_id).order_by(AssetHistory.issued_date.asc()).all()
                    if assets:
                        rows = []
                        for a in assets:
                            rows.append({
                                'Employee ID': a.employee_id,
                                'Asset Type': a.asset_type or '',
                                'Asset Number': a.asset_number or '',
                                'Issued Date (DD-MM-YYYY)': a.issued_date.strftime('%d-%m-%Y') if a.issued_date else '',
                                'Status': a.status or 'Issued',
                            })
                        asset_df = pd.DataFrame(rows)
                    else:
                        asset_df = pd.DataFrame(asset_data)
                else:
                    # All employees update - populate with all existing asset data
                    assets = db.query(AssetHistory).order_by(AssetHistory.employee_id.asc(), AssetHistory.issued_date.asc()).all()
                    if assets:
                        asset_data = {
                            'Employee ID': [a.employee_id for a in assets],
                            'Asset Type': [a.asset_type or '' for a in assets],
                            'Asset Number': [a.asset_number or '' for a in assets],
                            'Issued Date (DD-MM-YYYY)': [a.issued_date.strftime('%d-%m-%Y') if a.issued_date else '' for a in assets],
                            'Status': [a.status or 'Issued' for a in assets]
                        }
                        asset_df = pd.DataFrame(asset_data)
                    else:
                        asset_df = pd.DataFrame(asset_data)
            else:
                asset_df = pd.DataFrame(asset_data)
            asset_df.to_excel(writer, sheet_name='Asset Details', index=False)
            
            # Sheet 14: Instructions
            instructions_data = {
                'Instructions': [
                    '1. Fill in all required fields marked with *',
                    '2. For multiple entries (Family, Education, Experience, Address), add new rows as needed',
                    '3. Use DD-MM-YYYY format for all dates',
                    '4. For dropdown fields, use only the values specified in parentheses',
                    '5. Auto-generated fields will be calculated by the system',
                    '6. Do not modify column headers or sheet names',
                    '7. Leave empty cells for optional fields',
                    '8. Save as .xlsx format before uploading',
                    '',
                    'IMPORTANT - Employee ID Mapping:',
                    '- Each employee must have a UNIQUE Employee ID',
                    '- In multi-row sheets (Family, Education, Experience, Address, Emergency),',
                    '  copy the SAME Employee ID for all rows belonging to that employee',
                    '- Example: Employee "EMP001" has 2 education records and 3 family members:',
                    '  Education sheet: 2 rows both with Employee ID "EMP001"',
                    '  Family sheet: 3 rows all with Employee ID "EMP001"',
                    '',
                    'Required Fields:',
                    '- Employee ID (must be unique)',
                    '- First Name, Last Name',
                    '- DOJ (Date of Joining)',
                    '- Department, Designation',
                    '- Manager Name',
                    '- Official Contact Number',
                    '- Official Email ID',
                    '- Category',
                    '- Gender',
                    '- DOB (Date of Birth)',
                    '',
                    'Multiple Entry Sections (with Employee ID):',
                    '- Address Details: Add rows for Permanent, Temporary, Office addresses',
                    '  Each row must have the Employee ID of the employee',
                    '- Family Members: Add rows for each family member',
                    '  Each row must have the Employee ID of the employee',
                    '- Education Details: Add rows for each education record',
                    '  Each row must have the Employee ID of the employee',
                    '- Experience Details: Add rows for each work experience',
                    '  Each row must have the Employee ID of the employee',
                    '- Onboarding Details: One row per employee',
                    '  Row must include Employee ID and valid Client Name',
                    '- Asset Details: Add rows for each asset issued',
                    '  Each row must have the Employee ID of the employee',
                    '- Emergency Contacts: Exactly 2 rows (Contact 1 and Contact 2)',
                    '  Both rows must have the same Employee ID',
                    '',
                    'Example for 3 employees:',
                    'Employee Details: 3 rows (EMP001, EMP002, EMP003)',
                    'Education: 5 rows (2 for EMP001, 1 for EMP002, 2 for EMP003)',
                    'Experience: 4 rows (1 for EMP001, 2 for EMP002, 1 for EMP003)',
                    'Family: 6 rows (2 for EMP001, 2 for EMP002, 2 for EMP003)'
                ]
            }
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False)

            # NOW ADD DATA VALIDATIONS WHILE WRITER CONTEXT IS STILL OPEN
            try:
                wb = writer.book
                lists_ws = wb['Lists']
                # Hide the Lists sheet similar to candidates template
                lists_ws.sheet_state = 'hidden'

                # Build column index map for lists
                lists_headers = {cell.value: idx+1 for idx, cell in enumerate(lists_ws[1])}
                
                def list_range(header):
                    col_idx = lists_headers.get(header)
                    if not col_idx:
                        return None
                    col_letter = get_column_letter(col_idx)
                    max_row = max(lists_ws.max_row, 2)
                    return f"Lists!${col_letter}$2:${col_letter}${max_row}"

                def apply_list_validation(sheet_name, header_label, list_header):
                    if sheet_name not in wb.sheetnames:
                        return
                    ws = wb[sheet_name]
                    # locate column by header label
                    header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1])}
                    col_idx = header_map.get(header_label)
                    if not col_idx:
                        return
                    rng = list_range(list_header)
                    if not rng:
                        return
                    dv = DataValidation(type="list", formula1=rng, allow_blank=True)
                    ws.add_data_validation(dv)
                    # apply to many rows
                    col_letter = get_column_letter(col_idx)
                    dv.add(f"${col_letter}$2:${col_letter}$2000")

                # Apply validations to consolidated Employee Details sheet
                apply_list_validation('Employee Details', 'Department', 'Departments')
                apply_list_validation('Employee Details', 'Category', 'Categories')
                apply_list_validation('Employee Details', 'Excluded from Payroll', 'ExcludedFromPayroll')
                apply_list_validation('Employee Details', 'Title', 'Titles')
                apply_list_validation('Employee Details', 'Gender', 'Gender')
                apply_list_validation('Employee Details', 'Marital Status', 'MaritalStatus')
                apply_list_validation('Employee Details', 'Blood Group', 'BloodGroup')
                apply_list_validation('Employee Details', 'Job Type', 'JobType')
                # Address
                apply_list_validation('Address Details', 'Address Type', 'AddressType')
                # Family Members
                apply_list_validation('Family Members', 'Relation Type', 'RelationType')
                # Education
                apply_list_validation('Education Details', 'Type of Degree', 'DegreeType')
                # Experience
                apply_list_validation('Experience Details', 'Department', 'Departments')
                # Onboarding
                apply_list_validation('Onboarding Details', 'Client Name', 'Clients')
                apply_list_validation('Onboarding Details', 'Department', 'Departments')
                # Asset
                apply_list_validation('Asset Details', 'Asset Type', 'AssetType')
                apply_list_validation('Asset Details', 'Status', 'AssetStatus')
                
                # Set active sheet to Employee Details
                if 'Employee Details' in wb.sheetnames:
                    wb.active = wb.sheetnames.index('Employee Details')
                
            except Exception as e:
                # Do not fail template generation if validation wiring hits an edge-case
                print('Excel validation wiring error:', str(e))

        # Get the Excel file content
        output.seek(0)
        excel_content = output.getvalue()
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Employee_Data_Template_{timestamp}.xlsx"
        
        # Return the Excel file as a streaming response
        return StreamingResponse(
            io.BytesIO(excel_content),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel template: {str(e)}")