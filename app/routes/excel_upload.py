
# from fastapi import APIRouter, Depends
# from fastapi.responses import StreamingResponse
# from sqlalchemy.orm import Session
# from io import BytesIO
# from openpyxl import Workbook
# from openpyxl.worksheet.datavalidation import DataValidation
# from openpyxl.styles import numbers, Font, PatternFill

# from app.database import get_db
# from app.models import (
#     CTCStatus, Candidate, Department, Discussion, DiscussionStatusDB, FinalStatusDB, GenderDB, HRTeam, InterviewStatusDB, Job, ModeDB, OfferStatusDB, SecondInterviewTeam, StatusDB, RequisitionTypeDB, InterviewTeam,
#     TATeam  # Add your actual DB models here
# )

# router = APIRouter()

# @router.get("/download-template")
# def download_template(db: Session = Depends(get_db)):
#     # Fetch dropdown values from DB
#     current_statuses = [row.status for row in db.query(StatusDB).all()]
#     departments = [row.name for row in db.query(Department).all()]
#     final_statuses = [row.status for row in db.query(FinalStatusDB).all()]
#     genders = [row.gender for row in db.query(GenderDB).all()]
#     modes_of_work = [row.mode for row in db.query(ModeDB).all()]
#     associated_jobId = [row.job_title for row in db.query(Job).all()]
#     ctc_breakup_statuses = [row[0] for row in db.query(CTCStatus.status).distinct().all()]
#     current_offer_statuses = [row.status for row in db.query(OfferStatusDB).all()]
#     interview_statuses = [row.status for row in db.query(InterviewStatusDB).all()]
#     candidate_discussion_status =[row.status for row in db.query(DiscussionStatusDB).all()]

#     # Initialize workbook
#     wb = Workbook()
#     ws = wb.active
#     ws.title = "Candidate Upload Template"

#     # Define headers exactly as provided (61 headers)
#     headers = [
#         "Candidate Name", "Gender", "Email ID", "Mobile No",  "Date of Resume Received", 
#         "Department", "Associated with (Job ID)", "Current Location", "Application Date", 
#         "Skills", "Current Company", "Current designation", "Years of Experience", 
#         "Current Status", "Final Status", "Linked In Url", "Notice Period (In Days)", 
#         "Additional information of NPD", "Current Fixed CTC (In INR)", 
#         "Current Variable Pay (In INR)", "Expected Fixed CTC (In INR)", "Mode of Work", 
#         "Reason for job change", "TA Team Member", "TA comments", "L1 Interview Date", 
#         "L1 Interviewer Name", "L1 Status", "L1 Interview Feedback", "L2 Interview Date", 
#         "L2 Interviewer Name", "L2 Status", "L2 Interview Feedback", "HR Interview Date", 
#         "HR Interviewer Name", "HR Status", "Finalized CTC (In INR)", "HR Interview Feedback", 
#         "Current Address", "Permanent Address", "Expected Date of Joining", 
#         "CTC Breakup Status", "Current Offer Status", "Offer Status Date", 
#         "D1 Candidate Joining Status", "D1 Feedback", "D1 Done By", 
#         "D2 Candidate Joining Status", "D2 Feedback", "D2 Done By", 
#         "D3 Candidate Joining Status", "D3 Feedback", "D3 Done By", 
#         "D4 Candidate Joining Status", "D4 Feedback", "D4 Done By", 
#         "D5 Candidate Joining Status", "D5 Feedback", "D5 Done By", 
#         "D6 Candidate Joining Status", "D6 Feedback", "D6 Done By", 
#         "Employee Number", "Date of Joining"
#     ]

#     ws.append(headers)




#     # Map dropdowns using column letters
#     col_dropdown_map = {
#         "B": genders,                    # Gender
#         "F": departments,
#         "G":associated_jobId,                # Department
#         "N": current_statuses,           # Current Status
#         "O": final_statuses,             # Final Status
#         "V": modes_of_work,              # Mode of Work          
#         "AB": interview_statuses,        # L1 Status
#         "AF": interview_statuses,        # L2 Status
#         "AJ": interview_statuses,        # HR Status
#         "AO": ctc_breakup_statuses,      # CTC Breakup Status
#         "AP": current_offer_statuses,    # Current Offer Status
#         "AQ": current_offer_statuses,
#         "AS":candidate_discussion_status,
#         "AV":candidate_discussion_status,
#         "AY":candidate_discussion_status,



#     }

#     def safe_string(value):
#         try:
#             return str(value).replace('"', '""')
#         except:
#             return ""

#     # Add dropdown validations
#     for col_letter, values in col_dropdown_map.items():
#         if values:
#             safe_values = [safe_string(v) for v in values]
#             formula = f'"{",".join(safe_values)}"'

#             # Skip if formula is too long for Excel
#             if len(formula) > 255:
#                 continue  # Could log or use hidden sheet instead

#             dv = DataValidation(type="list", formula1=formula, allow_blank=True)
#             dv.error = "Please select a value from the dropdown list"
#             dv.errorTitle = "Invalid Selection"
#             ws.add_data_validation(dv)
#             dv.add(f"{col_letter}2:{col_letter}1000")

#     # Format specific columns
#     for row in range(2, 1001):
#         # Mobile number formatting (Column D)
#         ws[f"D{row}"].number_format = '0000000000'
        
#         dv = DataValidation(
#             type="textLength",
#             operator="equal",
#             formula1=10,
#             showErrorMessage=True,
#             errorTitle="Invalid Length",
#             error="Mobile number must be exactly 10 digits"
#         )

#         ws.add_data_validation(dv)
#         dv.add(f"D2:D1000")

#         # Date columns formatting
#         date_columns = ["E", "I", "Z", "AD", "AH", "AN", "AQ", "BI"]  # All date fields
#         for col in date_columns:
#             ws[f"{col}{row}"].number_format = 'DD-MM-YYYY'
        
#         # CTC columns formatting (currency in INR)
#         ctc_columns = ["S", "T", "U", "AK"]  # Current Fixed CTC, Variable Pay, Expected CTC, Finalized CTC
#         for col in ctc_columns:
#             ws[f"{col}{row}"].number_format = '₹#,##0'
        
#         # Notice Period formatting (numbers)
#         ws[f"Q{row}"].number_format = '0'  # Notice Period (In Days)
        
#         # Years of Experience formatting (decimal)
#         ws[f"M{row}"].number_format = '0.0'  # Years of Experience
        
#         # Employee Number formatting
#         ws[f"BH{row}"].number_format = '0'  # Employee Number

#     # Auto-adjust column widths for better readability
#     for column in ws.columns:
#         max_length = 0
#         column_letter = column[0].column_letter
#         for cell in column:
#             try:
#                 if len(str(cell.value)) > max_length:
#                     max_length = len(str(cell.value))
#             except:
#                 pass
#         adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
#         ws.column_dimensions[column_letter].width = max(adjusted_width, 12)  # Minimum width of 12

#     # Add styling to headers
#     header_font = Font(bold=True, color="FFFFFF")
#     header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
#     for cell in ws[1]:
#         cell.font = header_font
#         cell.fill = header_fill

#     # Freeze the header row for better navigation
#     ws.freeze_panes = "A2"

#     # Save workbook
#     stream = BytesIO()
#     wb.save(stream)
#     stream.seek(0)

#     return StreamingResponse(
#         stream,
#         media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
#         headers={"Content-Disposition": "attachment; filename=CandidateUploadTemplate.xlsx"}
#     )



from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from io import BytesIO
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import numbers, Font, PatternFill
import openpyxl.comments

from app.database import get_db
from app.models import (
    CTCStatus, Candidate, Department, Discussion, DiscussionStatusDB, FinalStatusDB, GenderDB, HRTeam, InterviewStatusDB, Job, ModeDB, OfferStatusDB, SecondInterviewTeam, StatusDB, RequisitionTypeDB, InterviewTeam,
    TATeam, TalentAcquisitionTeam  # Add TalentAcquisitionTeam for TA team members
)

router = APIRouter()

@router.get("/download-template")
def download_template(db: Session = Depends(get_db)):
    # Fetch dropdown values from DB
    current_statuses = [row.status for row in db.query(StatusDB).all()]
    departments = [row.name for row in db.query(Department).all()]
    # Removed: final_statuses dropdown as final_status column is not present in downloadable template
    # final_statuses = [row.status for row in db.query(FinalStatusDB).all()]
    genders = [row.gender for row in db.query(GenderDB).all()]
    modes_of_work = [row.mode for row in db.query(ModeDB).all()]
    associated_jobId = [row.job_title for row in db.query(Job).all()]
    # Get offer statuses from the offer_statuses table (same as /candidates/offer_status/all endpoint)
    current_offer_statuses = [row.status for row in db.query(OfferStatusDB).all()]
    # Use the same offer statuses for ctc_breakup_status dropdown
    ctc_breakup_statuses = current_offer_statuses
    interview_statuses = [row.status for row in db.query(InterviewStatusDB).all()]
    candidate_discussion_status = [row.status for row in db.query(DiscussionStatusDB).all()]
    
    # Fetch TA team members for dropdown and convert to name.lastname format
    ta_team_members = []
    ta_teams = db.query(TalentAcquisitionTeam).all()
    for team in ta_teams:
        for member in team.team_members:
            # Convert "Amarnath Dursheti" to "amarnath.dursheti"
            if member and ' ' in member:
                parts = member.lower().split()
                if len(parts) >= 2:
                    formatted_name = f"{parts[0]}.{parts[-1]}"
                    ta_team_members.append(formatted_name)
            else:
                ta_team_members.append(member.lower() if member else member)

    # Build interviewer dropdown items per team.
    # For each team, if there are multiple emails in the cell, join the local parts with commas (no spaces).
    # If there is a single email, use just its local part. Example:
    # "nrendra.kumar@vaics...,jai.kumar@vaics..." -> "nrendra.kumar,jai.kumar"
    def build_interviewer_values_from_teams(teams):
        interviewer_values = []
        for team in teams or []:
            try:
                raw_emails = getattr(team, 'team_emails', None)
                if not raw_emails:
                    continue

                # Normalize to a list of email strings
                if isinstance(raw_emails, list):
                    email_list = []
                    for entry in raw_emails:
                        if entry is None:
                            continue
                        # entry can also be a comma-separated string
                        parts = [e.strip() for e in str(entry).split(',') if e and e.strip()]
                        email_list.extend(parts)
                else:
                    email_list = [e.strip() for e in str(raw_emails).split(',') if e and e.strip()]

                # Extract local parts and join with commas (no spaces)
                local_parts = []
                for email in email_list:
                    if '@' in email:
                        local = email.split('@')[0].strip()
                        if local:
                            local_parts.append(local)

                if local_parts:
                    interviewer_values.append(','.join(local_parts))
            except Exception:
                continue
        # Keep order stable but unique
        seen = set()
        deduped = []
        for item in interviewer_values:
            if item not in seen:
                seen.add(item)
                deduped.append(item)
        return deduped

    # L1 interviewers from interview_teams (per-team combined local parts)
    l1_interviewer_names = build_interviewer_values_from_teams(db.query(InterviewTeam).all())

    # L2 interviewers from second_interview_teams (per-team combined local parts)
    l2_interviewer_names = build_interviewer_values_from_teams(db.query(SecondInterviewTeam).all())

    # HR interviewers from hr_teams (per-team combined local parts)
    hr_interviewer_names = build_interviewer_values_from_teams(db.query(HRTeam).all())

    # Discussion done_by dropdowns from TalentAcquisitionTeam members (names, not emails)
    discussion_team_members = []
    ta_teams = db.query(TalentAcquisitionTeam).all()
    for team in ta_teams:
        if getattr(team, 'team_members', None):
            discussion_team_members.extend(team.team_members)

    # Initialize workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidate Upload Template"

    # Define headers with corrected order and names
    headers = [
        "candidate_name",      # Column A (1)
        "email_id",           # Column B (2)
        "mobile_no",          # Column C (3)
        "pan_card_no",        # Column D (4)
        "date_of_resume_received", # Column E (5) - MOVED BEFORE associated_job_id
        "department",         # Column F (6)
        "associated_job_id",  # Column G (7) - MOVED AFTER date_of_resume_received
        "application_date",   # Column H (8)
        "skills_set",         # Column I (9)
        "current_company",    # Column J (10)
        "current_designation", # Column K (11)
        "gender",             # Column L (12)
        "years_of_experience", # Column M (13)
        "current_status",     # Column N (14)
        "final_status",       # Column O (15) - ADDED final_status field
        "notice_period",      # Column P (16)
        "current_location",   # Column Q (17)
        "additional_information_npd", # Column R (18)
        "current_fixed_ctc(IN INR)", # Column S (19)
        "current_variable_pays (IN INR)", # Column T (20)
        "expected_fixed_ctc (IN INR)", # Column U (21)
        "mode_of_work",       # Column V (22)
        "reason_for_job_change", # Column W (23)
        "ta_team",            # Column X (24)
        "ta_comments",        # Column Y (25)
        "linkedin_url",       # Column Z (26)
        "l1_interview_date",  # Column AA (27)
        "l1_interviewer_name", # Column AB (28)
        "l1_status",          # Column AC (29)
        "l1_feedback",        # Column AD (30) - FIXED: was l1_interview_feedback
        "l2_interview_date",  # Column AE (31)
        "l2_interviewer_name", # Column AF (32)
        "l2_status",          # Column AG (33)
        "l2_feedback",        # Column AH (34) - FIXED: was l2_interview_feedback
        "hr_interview_date",  # Column AI (35)
        "hr_interviewer_name", # Column AJ (36)
        "hr_status",          # Column AK (37)
        "hr_feedback",        # Column AL (38) - FIXED: was hr_interview_feedback
        "Finalized CTC (In INR)", # Column AM (39) - RENAMED: was fixed_ctc_lakhs
        "designation",        # Column AN (40)
        "Offered CTC (in INR)", # Column AO (41) - RENAMED: was ctc
        "ctc_breakup_status", # Column AP (42)
        "current_address",    # Column AQ (43)
        "permanent_address",  # Column AR (44)
        "expected_date_of_joining", # Column AS (45)
        "offer_status",       # Column AT (46) - RENAMED: was offer_letter_status
        "discussion1_status", # Column AU (47)
        "discussion1_done_by", # Column AV (48)
        "discussion1_notes",  # Column AW (49)
        "discussion2_status", # Column AX (50)
        "discussion2_done_by", # Column AY (51)
        "discussion2_notes",  # Column AZ (52)
        "discussion3_status", # Column BA (53)
        "discussion3_done_by", # Column BB (54)
        "discussion3_notes",  # Column BC (55)
        "discussion4_status", # Column BD (56)
        "discussion4_done_by", # Column BE (57)
        "discussion4_notes",  # Column BF (58)
        "discussion5_status", # Column BG (59)
        "discussion5_done_by", # Column BH (60)
        "discussion5_notes",  # Column BI (61)
        "discussion6_status", # Column BJ (62)
        "discussion6_done_by", # Column BK (63)
        "discussion6_notes"   # Column BL (64)
    ]

    ws.append(headers)

    # Add a sample row to show expected format
    sample_row = [
        "John Doe",                    # candidate_name
        "john.doe@example.com",        # email_id
        "9876543210",                  # mobile_no - EXAMPLE: 10 digits
        "ABCDE1234F",                  # pan_card_no - EXAMPLE: 10 characters
        "15-01-2024",                  # date_of_resume_received (dd-mm-yyyy format)
        "Engineering",                  # department
        "JOB123",                      # associated_job_id
        "16-01-2024",                  # application_date (dd-mm-yyyy format)
        "Python, React, SQL",          # skills_set
        "ABC Corp",                    # current_company
        "Senior Developer",            # current_designation
        "Male",                        # gender
        "5.5",                         # years_of_experience
        "Screening",                   # current_status
        # Removed final_status sample value as column will be deleted from template
        "30",                          # notice_period
        "New York",                    # current_location
        "Career growth opportunity",   # additional_information_npd
        "100000",                      # current_fixed_ctc
        "20000",                       # current_variable_pay
        "150000",                      # expected_fixed_ctc
        "Remote",                      # mode_of_work
        "Career Growth",               # reason_for_job_change
        "Engineering Team",            # ta_team
        "Strong candidate",            # ta_comments
        "https://linkedin.com/in/johndoe", # linkedin_url
        "20-01-2024",                 # l1_interview_date (dd-mm-yyyy format)
        "john.interviewer",            # l1_interviewer_name
        "Passed",                      # l1_status
        "Good technical skills",       # l1_feedback
        "25-01-2024",                 # l2_interview_date (dd-mm-yyyy format)
        "jane.interviewer",            # l2_interviewer_name
        "Passed",                      # l2_status
        "Excellent communication",     # l2_feedback
        "30-01-2024",                 # hr_interview_date (dd-mm-yyyy format)
        "hr.interviewer",              # hr_interviewer_name
        "Passed",                      # hr_status
        "Good cultural fit",           # hr_feedback
        "150000",                      # finalized_ctc
        "Senior Developer",            # designation
        "160000",                      # offered_ctc
        "Pending",                     # ctc_breakup_status
        "123 Main St, New York",      # current_address
        "123 Main St, New York",      # permanent_address
        "01-03-2024",                 # expected_date_of_joining (dd-mm-yyyy format)
        "Pending",                     # offer_status
        "Pending",                     # discussion1_status
        "John Manager",                # discussion1_done_by
        "Initial discussion completed", # discussion1_notes
        "Pending",                     # discussion2_status
        "Jane Manager",                # discussion2_done_by
        "Second discussion scheduled", # discussion2_notes
        "Pending",                     # discussion3_status
        "HR Manager",                  # discussion3_done_by
        "HR discussion pending",       # discussion3_notes
        "Pending",                     # discussion4_status
        "Director",                    # discussion4_done_by
        "Director approval pending",   # discussion4_notes
        "Pending",                     # discussion5_status
        "CEO",                         # discussion5_done_by
        "CEO approval pending",        # discussion5_notes
        "Pending",                     # discussion6_status
        "Board",                       # discussion6_done_by
        "Board approval pending"       # discussion6_notes
    ]
    ws.append(sample_row)

    # (Removed) Previously hid 'final_status' column (Column O)

    def safe_string(value):
        try:
            return str(value).replace('"', '""')
        except:
            return ""

    # CORRECTED dropdown mappings for the new column positions
    col_dropdown_map = {
        "L": genders,                    # gender (Column L - 12th column)
        "F": departments,                # department (Column F - 6th column)
        "G": associated_jobId,           # associated_job_id (Column G - 7th column)
        "N": current_statuses,           # current_status (Column N - 14th column)
        # Removed final_status dropdown to prevent shift onto notice_period after column deletion
        "V": modes_of_work,              # mode_of_work (Column V - 22nd column)
        "X": ta_team_members,            # ta_team (Column X - 24th column) - ADDED TA team dropdown
        # For interviewer name columns (AB, AF, AJ) we will use hidden-sheet ranges to allow commas inside items
        "AB": [],                        # placeholder; handled via hidden range
        "AC": interview_statuses,        # l1_status (Column AC - 29th column)
        "AF": [],                        # placeholder; handled via hidden range
        "AG": interview_statuses,        # l2_status (Column AG - 33rd column)
        "AJ": [],                        # placeholder; handled via hidden range
        "AK": interview_statuses,        # hr_status (Column AK - 37th column)
        "AP": current_offer_statuses,    # ctc_breakup_status (Column AP - 42nd column)
        "AT": current_offer_statuses,    # offer_status (Column AT - 46th column) - RENAMED
        "AU": candidate_discussion_status, # discussion1_status (Column AU - 47th column)
        "AV": discussion_team_members,     # discussion1_done_by (Column AV - 48th column)
        "AX": candidate_discussion_status, # discussion2_status (Column AX - 51st column)
        "AY": discussion_team_members,     # discussion2_done_by (Column AY - 52nd column)
        "BA": candidate_discussion_status, # discussion3_status (Column BA - 55th column)
        "BB": discussion_team_members,     # discussion3_done_by (Column BB - 56th column)
        "BD": candidate_discussion_status, # discussion4_status (Column BD - 59th column)
        "BE": discussion_team_members,     # discussion4_done_by (Column BE - 60th column)
        "BG": candidate_discussion_status, # discussion5_status (Column BG - 63rd column)
        "BH": discussion_team_members,     # discussion5_done_by (Column BH - 64th column)
        "BJ": candidate_discussion_status, # discussion6_status (Column BJ - 67th column)
        "BK": discussion_team_members,     # discussion6_done_by (Column BK - 68th column)
    }

    # Build a hidden sheet for lists that must preserve commas inside items
    lists_ws = wb.create_sheet(title="Dropdowns")
    # L1 interviewer names in column A
    for idx, val in enumerate(l1_interviewer_names, start=1):
        lists_ws.cell(row=idx, column=1, value=val)
    l1_range = f"=Dropdowns!$A$1:$A${max(1, len(l1_interviewer_names))}"
    # L2 interviewer names in column B
    for idx, val in enumerate(l2_interviewer_names, start=1):
        lists_ws.cell(row=idx, column=2, value=val)
    l2_range = f"=Dropdowns!$B$1:$B${max(1, len(l2_interviewer_names))}"
    # HR interviewer names in column C
    for idx, val in enumerate(hr_interviewer_names, start=1):
        lists_ws.cell(row=idx, column=3, value=val)
    hr_range = f"=Dropdowns!$C$1:$C${max(1, len(hr_interviewer_names))}"
    # Hide the lists sheet
    lists_ws.sheet_state = 'hidden'

    # Add dropdown validations
    for col_letter, values in col_dropdown_map.items():
        # Interviewer name columns handled via hidden ranges to allow commas within items
        if col_letter == "AB":
            dv = DataValidation(type="list", formula1=l1_range, allow_blank=True)
            dv.error = "Please select a value from the dropdown list"
            dv.errorTitle = "Invalid Selection"
            ws.add_data_validation(dv)
            dv.add("AB2:AB1000")
            continue
        if col_letter == "AF":
            dv = DataValidation(type="list", formula1=l2_range, allow_blank=True)
            dv.error = "Please select a value from the dropdown list"
            dv.errorTitle = "Invalid Selection"
            ws.add_data_validation(dv)
            dv.add("AF2:AF1000")
            continue
        if col_letter == "AJ":
            dv = DataValidation(type="list", formula1=hr_range, allow_blank=True)
            dv.error = "Please select a value from the dropdown list"
            dv.errorTitle = "Invalid Selection"
            ws.add_data_validation(dv)
            dv.add("AJ2:AJ1000")
            continue

        if values:
            safe_values = [safe_string(v) for v in values]
            formula = f'"{",".join(safe_values)}"'

            # Skip if formula is too long for Excel
            if len(formula) > 255:
                continue  # Could log or use hidden sheet instead

            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
            dv.error = "Please select a value from the dropdown list"
            dv.errorTitle = "Invalid Selection"
            ws.add_data_validation(dv)
            dv.add(f"{col_letter}2:{col_letter}1000")

    # Apply simple text formatting for date columns to avoid Excel interpretation issues
    date_column_letters = ["E", "H", "AA", "AE", "AI", "AS"]  # date columns
    
    for col_letter in date_column_letters:
        # Set up text formatting for these columns to avoid date interpretation issues
        for row in range(2, 1001):
            cell = ws[f"{col_letter}{row}"]
            # Use text format to prevent Excel from auto-interpreting dates
            cell.number_format = '@'
            
            # Add text length validation for date format (dd-mm-yyyy = 10 characters)
            date_dv = DataValidation(
                type="textLength",
                operator="equal",
                formula1=10,
                showErrorMessage=True,
                errorTitle="Invalid Date Format",
                error="Please enter date in dd-mm-yyyy format (e.g., 15-08-2025)"
            )
            ws.add_data_validation(date_dv)
            date_dv.add(f"{col_letter}{row}")
            
            # Add a comment to the first date field to emphasize the format
            if row == 2 and col_letter == "E":
                ws[f"{col_letter}{row}"].comment = openpyxl.comments.Comment(
                    "Enter date as: dd-mm-yyyy (e.g., 15-08-2025)\nThis will be processed as text to avoid Excel date issues",
                    "HRMS System"
                )

    # Format specific columns
    for row in range(2, 1001):
        
        # Mobile number formatting (Column C) - Use text format to preserve exact input
        ws[f"C{row}"].number_format = '@'
        
        # Mobile number validation (10 digits)
        mobile_dv = DataValidation(
            type="textLength",
            operator="equal",
            formula1=10,
            showErrorMessage=True,
            errorTitle="Invalid Length",
            error="Mobile number must be exactly 10 digits"
        )
        ws.add_data_validation(mobile_dv)
        mobile_dv.add(f"C2:C1000")
        
        # Add a note in the first row to help users
        if row == 2:
            ws[f"C{row}"].comment = openpyxl.comments.Comment(
                "Enter exactly 10 digits (e.g., 9876543210)\nNo spaces, dashes, or country codes\nThis field uses TEXT format to preserve your input exactly",
                "HRMS System"
            )

        # PAN Card Number formatting and validation (Column D)
        ws[f"D{row}"].number_format = '@'  # Text format
        
        # PAN validation (10 characters: 5 letters + 4 digits + 1 letter)
        pan_dv = DataValidation(
            type="textLength",
            operator="equal",
            formula1=10,
            showErrorMessage=True,
            errorTitle="Invalid PAN Format",
            error="PAN must be exactly 10 characters (e.g., ABCDE1234F)"
        )
        ws.add_data_validation(pan_dv)
        pan_dv.add(f"D2:D1000")

        # Date columns formatting is now handled above in the date_column_letters loop
        
        # Notice Period formatting (numbers) - Column P
        ws[f"P{row}"].number_format = '0'
        
        # Years of Experience formatting (decimal) - Column M
        ws[f"M{row}"].number_format = '0.0'
        
        # CTC columns formatting (currency in INR)
        ctc_columns = ["S", "T", "U", "AM", "AO"]  # Current Fixed CTC, Variable Pay, Expected CTC, Finalized CTC, Offered CTC
        for col in ctc_columns:
            ws[f"{col}{row}"].number_format = '0'  # Simple number format without currency symbol

    # Auto-adjust column widths for better readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)  # Minimum width of 12

    # Add styling to headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
    
    # Style the sample row to make it clear it's an example
    sample_font = Font(bold=True, color="000000")
    sample_fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
    
    for cell in ws[2]:
        cell.font = sample_font
        cell.fill = sample_fill

    # Freeze the header row for better navigation
    ws.freeze_panes = "A2"

    # Add an instructions worksheet
    instructions_ws = wb.create_sheet("Instructions", 0)  # Insert as first sheet
    instructions_ws.append(["HRMS Candidate Upload Template - Instructions"])
    instructions_ws.append([])
    instructions_ws.append(["Sample Data:"])
    instructions_ws.append(["• Row 2 contains sample data to show the expected format"])
    instructions_ws.append(["• You can delete this row or use it as a reference"])
    instructions_ws.append([])
    instructions_ws.append(["Date Fields:"])
    instructions_ws.append(["• IMPORTANT: Always use dd-mm-yyyy format (e.g., 15-08-2025)"])
    instructions_ws.append(["• Do NOT use mm-dd-yyyy format (e.g., 08-15-2025) as it will be interpreted incorrectly"])
    instructions_ws.append(["• Common mistake: 11-08-2025 means 11th August 2025, NOT November 8th 2025"])
    instructions_ws.append(["• You can enter dates manually: 15-8-2025, 15/8/2025, 15-08-2025"])
    instructions_ws.append(["• Or use the calendar picker by clicking on the cell"])
    instructions_ws.append(["• The template enforces dd-mm-yyyy format to prevent date interpretation errors"])
    instructions_ws.append([])
    instructions_ws.append(["Skills Field:"])
    instructions_ws.append(["• Enter skills separated by commas: Python, React, Node.js"])
    instructions_ws.append(["• Both {skill1, skill2} and skill1, skill2 formats are supported"])
    instructions_ws.append([])
    instructions_ws.append(["Mobile Number Field:"])
    instructions_ws.append(["• Enter exactly 10 digits: 9876543210"])
    instructions_ws.append(["• Do not include spaces, dashes, or country codes"])
    instructions_ws.append(["• Leading zeros will be preserved"])
    instructions_ws.append([])
    instructions_ws.append(["IMPORTANT NOTES:"])
    instructions_ws.append(["• Dates must be in dd-mm-yyyy format to avoid interpretation errors"])
    instructions_ws.append(["• Mobile numbers use text format to preserve exact input"])
    instructions_ws.append(["• Sample row shows correct format for all fields"])
    
    # Style the instructions
    instructions_ws['A1'].font = Font(bold=True, size=14)
    instructions_ws['A3'].font = Font(bold=True)
    instructions_ws['A8'].font = Font(bold=True)
    instructions_ws['A12'].font = Font(bold=True)
    instructions_ws['A16'].font = Font(bold=True)
    instructions_ws['A20'].font = Font(bold=True)
    instructions_ws['A24'].font = Font(bold=True)
    
    # Set the main data sheet as active
    wb.active = ws

    # Remove 'final_status' column (Column O) from the downloadable template
    # This ensures the column is not present in the final Excel file
    ws.delete_cols(15, 1)

    # Save workbook
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=CandidateUploadTemplate.xlsx"}
    )